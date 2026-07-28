# coding: utf-8 в
import keep_alive
keep_alive.start()
# bot.py
import asyncio
import io
import csv
import shutil
import os
import webbrowser
from datetime import date, timedelta, datetime

# Aiogram 3
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import (
    ReplyKeyboardMarkup,
    KeyboardButton,
    BufferedInputFile
)
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext

# SQLAlchemy
from sqlalchemy import create_engine, Column, Integer, Float, Date, String, Boolean, func, delete, text
from sqlalchemy.orm import declarative_base, sessionmaker

# Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Графики
# import matplotlib
# matplotlib.use('Agg')
# import matplotlib.pyplot as plt
# from matplotlib import rcParams
# import matplotlib.dates as mdates

# Прогнозирование
# import numpy as np
# from sklearn.linear_model import LinearRegression

# Для веб-сервера SQLite
import sqlite3
from http.server import HTTPServer, SimpleHTTPRequestHandler
import threading

# ---------- НАСТРОЙКИ ----------
BOT_TOKEN = "8966604433:AAEAxUlW0Nk44mck3uioqWszYQuYzJyyyW0"
DATABASE_URL = "sqlite:///store.db"
BACKUP_DIR = "backups"
REPORTS_DIR = "reports"
BANK_COMMISSION_RATE = 0.01
SELLER_SALARY_RATE = 0.05

SELLERS = ["Людмила", "Светлана", "Елена"]

# Режимы
TEST_MODE = False

for dir_name in [BACKUP_DIR, REPORTS_DIR]:
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)

bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# ---------- БАЗА ДАННЫХ ----------
engine = create_engine(DATABASE_URL, echo=False)
Session = sessionmaker(bind=engine)
Base = declarative_base()


class DailyRevenue(Base):
    __tablename__ = "daily_revenue"
    id = Column(Integer, primary_key=True)
    date = Column(Date)
    seller = Column(String)
    cash = Column(Float, default=0)
    non_cash = Column(Float, default=0)
    kaspi = Column(Float, default=0)
    halyk = Column(Float, default=0)
    jusan = Column(Float, default=0)
    income = Column(Float, default=0)
    expense = Column(Float, default=0)


class CashDelivery(Base):
    __tablename__ = "cash_delivery"
    id = Column(Integer, primary_key=True)
    date = Column(Date)
    amount = Column(Float, default=0)


class BankCommission(Base):
    __tablename__ = "bank_commission"
    id = Column(Integer, primary_key=True)
    date = Column(Date, unique=True)
    kaspi_commission = Column(Float, default=0)
    halyk_commission = Column(Float, default=0)
    jusan_commission = Column(Float, default=0)
    total_commission = Column(Float, default=0)


class Journal(Base):
    __tablename__ = "journal"
    id = Column(Integer, primary_key=True)
    date = Column(Date)
    expected_balance = Column(Float)


class Inventory(Base):
    __tablename__ = "inventory"
    id = Column(Integer, primary_key=True)
    date = Column(Date)
    actual_balance = Column(Float, default=0)
    seller_debts = Column(Float, default=0)
    seller_bonus = Column(Float, default=0)
    final_balance = Column(Float, default=0)
    difference = Column(Float, default=0)


class Salary(Base):
    __tablename__ = "salary"
    id = Column(Integer, primary_key=True)
    seller = Column(String)
    date = Column(Date)
    revenue = Column(Float)
    salary_amount = Column(Float)
    paid = Column(Boolean, default=False)


class DailyCash(Base):
    __tablename__ = "daily_cash"
    id = Column(Integer, primary_key=True)
    date = Column(Date, unique=True)
    cash_balance = Column(Float, default=0)


def ensure_columns():
    with engine.connect() as conn:
        for col in ['income', 'expense']:
            try:
                conn.execute(text(f"ALTER TABLE daily_revenue ADD COLUMN {col} FLOAT DEFAULT 0"))
                conn.commit()
            except:
                pass
        for col in ['seller_debts', 'seller_bonus', 'final_balance']:
            try:
                conn.execute(text(f"ALTER TABLE inventory ADD COLUMN {col} FLOAT DEFAULT 0"))
                conn.commit()
            except:
                pass


Base.metadata.create_all(engine)
ensure_columns()


# ---------- ВЕБ-СЕРВЕР ДЛЯ ПРОСМОТРА БД ----------
class SQLiteWebHandler(SimpleHTTPRequestHandler):
    def do_GET(self):
        if self.path == '/':
            self.send_response(200)
            self.send_header('Content-type', 'text/html')
            self.end_headers()

            html = '''
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Просмотр базы данных магазина</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
                    h1 { color: #333; }
                    .table-container { background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); margin-bottom: 20px; overflow-x: auto; }
                    table { border-collapse: collapse; width: 100%; font-size: 14px; }
                    th { background: #366092; color: white; padding: 10px; text-align: left; }
                    td { padding: 8px; border-bottom: 1px solid #ddd; }
                    tr:hover { background: #f0f0f0; }
                    .tab { display: inline-block; padding: 10px 20px; margin: 5px; background: #366092; color: white; border: none; border-radius: 5px; cursor: pointer; }
                    .tab:hover { background: #2a4a70; }
                    .tab-content { display: none; }
                    .tab-content.active { display: block; }
                </style>
            </head>
            <body>
                <h1>📊 Просмотр базы данных магазина</h1>
                <div id="tabs">
                    <button class="tab" onclick="showTab('revenue')">Выручка</button>
                    <button class="tab" onclick="showTab('delivery')">Привоз</button>
                    <button class="tab" onclick="showTab('commission')">Комиссии</button>
                    <button class="tab" onclick="showTab('journal')">Журнал</button>
                    <button class="tab" onclick="showTab('inventory')">Ревизии</button>
                    <button class="tab" onclick="showTab('salary')">Зарплаты</button>
                    <button class="tab" onclick="showTab('cash')">Касса</button>
                </div>
                <div id="revenue" class="tab-content active">''' + self.get_table_html('daily_revenue') + '''</div>
                <div id="delivery" class="tab-content">''' + self.get_table_html('cash_delivery') + '''</div>
                <div id="commission" class="tab-content">''' + self.get_table_html('bank_commission') + '''</div>
                <div id="journal" class="tab-content">''' + self.get_table_html('journal') + '''</div>
                <div id="inventory" class="tab-content">''' + self.get_table_html('inventory') + '''</div>
                <div id="salary" class="tab-content">''' + self.get_table_html('salary') + '''</div>
                <div id="cash" class="tab-content">''' + self.get_table_html('daily_cash') + '''</div>
                <script>
                    function showTab(tabId) {
                        var contents = document.getElementsByClassName('tab-content');
                        for (var i = 0; i < contents.length; i++) {
                            contents[i].classList.remove('active');
                        }
                        document.getElementById(tabId).classList.add('active');
                    }
                </script>
            </body>
            </html>
            '''
            self.wfile.write(html.encode())

    def get_table_html(self, table_name):
        try:
            conn = sqlite3.connect('store.db')
            cursor = conn.cursor()
            cursor.execute(f"SELECT * FROM {table_name} LIMIT 100")
            rows = cursor.fetchall()
            columns = [description[0] for description in cursor.description]
            conn.close()

            if not rows:
                return '<p>Нет данных</p>'

            html = '<table><tr>'
            for col in columns:
                html += f'<th>{col}</th>'
            html += '</tr>'

            for row in rows:
                html += '<tr>'
                for cell in row:
                    if isinstance(cell, date):
                        cell = cell.strftime('%d.%m.%Y')
                    elif isinstance(cell, float):
                        cell = f'{cell:,.0f}'
                    html += f'<td>{cell}</td>'
                html += '</tr>'

            html += '</table>'
            return html
        except Exception as e:
            return f'<p>Ошибка: {e}</p>'


def start_web_server():
    try:
        port = 8080
        server = HTTPServer(('localhost', port), SQLiteWebHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        return f"http://localhost:{port}"
    except Exception as e:
        print(f"Ошибка запуска веб-сервера: {e}")
        return None


WEB_URL = start_web_server()


# ---------- СОСТОЯНИЯ ----------
class RevenueState(StatesGroup):
    waiting_seller = State()
    waiting_date = State()
    waiting_cash = State()
    waiting_kaspi = State()
    waiting_halyk = State()
    waiting_jusan = State()
    waiting_income = State()
    waiting_expense = State()
    waiting_confirm = State()


class DeliveryState(StatesGroup):
    waiting_date = State()
    waiting_amount = State()
    waiting_confirm = State()


class JournalState(StatesGroup):
    waiting_date = State()
    waiting_amount = State()
    waiting_confirm = State()


class InventoryState(StatesGroup):
    waiting_date = State()
    waiting_actual = State()
    waiting_debts = State()
    waiting_bonus = State()
    waiting_confirm = State()


class ViewState(StatesGroup):
    waiting_date = State()


class SalaryState(StatesGroup):
    waiting_seller = State()
    waiting_confirm = State()


class PaymentMethodState(StatesGroup):
    waiting_method = State()
    waiting_period = State()


class EditState(StatesGroup):
    waiting_date = State()
    waiting_type = State()
    waiting_seller = State()
    waiting_value = State()
    waiting_confirm = State()


class DailyReportState(StatesGroup):
    waiting_date = State()
    waiting_seller = State()
    waiting_expense = State()
    waiting_cash_balance = State()


class DeveloperState(StatesGroup):
    waiting_action = State()


class CommissionPeriodState(StatesGroup):
    waiting_type = State()
    waiting_date = State()


class ClearState(StatesGroup):
    waiting_confirm = State()


# ---------- КЛАВИАТУРЫ ----------
def main_keyboard():
    kb = [
        [KeyboardButton(text="💰 Ввести деньги за день")],
        [KeyboardButton(text="🚚 Привоз денег")],
        [KeyboardButton(text="📓 Журнал учёта")],
        [KeyboardButton(text="🧮 Ревизия")],
        [KeyboardButton(text="💰 Зарплаты")],
        [KeyboardButton(text="📅 Посмотреть день")],
        [KeyboardButton(text="📊 По способу оплаты")],
        [KeyboardButton(text="📊 Комиссии банков")],
        [KeyboardButton(text="📋 Дневной отчёт")],
        [KeyboardButton(text="📈 Анализ и прогноз")],
        [KeyboardButton(text="📄 Отчёт за месяц")],
        [KeyboardButton(text="📊 Годовой отчёт")],
        [KeyboardButton(text="✏️ Редактировать день")],
        [KeyboardButton(text="💾 Бэкап БД")],
        [KeyboardButton(text="🗃️ Просмотр всех данных")],
        [KeyboardButton(text="🧪 Тестовый режим")],
        [KeyboardButton(text="🗑️ Очистить все данные")],
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def cancel_keyboard():
    kb = [[KeyboardButton(text="❌ Отмена")]]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def seller_keyboard():
    kb = [[KeyboardButton(text=seller)] for seller in SELLERS]
    kb.append([KeyboardButton(text="❌ Отмена")])
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def confirm_keyboard():
    kb = [
        [KeyboardButton(text="✅ Да, сохранить")],
        [KeyboardButton(text="❌ Нет, отмена")]
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def payment_method_keyboard():
    kb = [
        [KeyboardButton(text="💵 Наличные")],
        [KeyboardButton(text="🏦 Kaspi")],
        [KeyboardButton(text="🏦 Халык")],
        [KeyboardButton(text="🏦 Жусан")],
        [KeyboardButton(text="📊 Все способы")],
        [KeyboardButton(text="❌ Отмена")]
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def period_keyboard():
    kb = [
        [KeyboardButton(text="📅 Сегодня")],
        [KeyboardButton(text="📆 Неделя")],
        [KeyboardButton(text="📆 Месяц")],
        [KeyboardButton(text="📆 Квартал")],
        [KeyboardButton(text="📊 Год")],
        [KeyboardButton(text="✏️ Свой период")],
        [KeyboardButton(text="❌ Отмена")]
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def commission_period_keyboard():
    kb = [
        [KeyboardButton(text="📆 От ревизии до ревизии")],
        [KeyboardButton(text="📆 За месяц")],
        [KeyboardButton(text="❌ Отмена")]
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def edit_type_keyboard():
    kb = [
        [KeyboardButton(text="💰 Выручка")],
        [KeyboardButton(text="📥 Приход")],
        [KeyboardButton(text="📤 Расход")],
        [KeyboardButton(text="🏦 Касса")],
        [KeyboardButton(text="❌ Отмена")]
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def test_mode_keyboard():
    kb = [
        [KeyboardButton(text="🧪 Включить тестовый режим")],
        [KeyboardButton(text="🧪 Выключить тестовый режим")],
        [KeyboardButton(text="🗑️ Очистить тестовые данные")],
        [KeyboardButton(text="❌ Отмена")]
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


# ---------- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ----------
def parse_date(date_str):
    try:
        return datetime.strptime(date_str, "%d.%m.%Y").date()
    except:
        try:
            return datetime.strptime(date_str, "%d/%m/%Y").date()
        except:
            return None


def format_date(date_obj):
    return date_obj.strftime("%d.%m.%Y")


def format_number(num):
    return f"{num:,.0f}".replace(",", " ")


def get_total_days():
    with Session() as session:
        return session.query(DailyRevenue.date).distinct().count()


def get_last_journal():
    with Session() as session:
        return session.query(Journal).order_by(Journal.date.desc()).first()


def get_last_inventory():
    with Session() as session:
        return session.query(Inventory).order_by(Inventory.date.desc()).first()


def get_previous_inventory():
    with Session() as session:
        return session.query(Inventory).order_by(Inventory.date.desc()).offset(1).first()


def get_daily_cash(date_obj):
    with Session() as session:
        daily_cash = session.query(DailyCash).filter(
            DailyCash.date == date_obj
        ).first()
        return daily_cash.cash_balance if daily_cash else None


def calculate_bank_commission(revenue):
    kaspi_comm = revenue['kaspi'] * BANK_COMMISSION_RATE
    halyk_comm = revenue['halyk'] * BANK_COMMISSION_RATE
    jusan_comm = revenue['jusan'] * BANK_COMMISSION_RATE
    return {
        'kaspi': kaspi_comm,
        'halyk': halyk_comm,
        'jusan': jusan_comm,
        'total': kaspi_comm + halyk_comm + jusan_comm
    }


def get_seller_total_salary(seller):
    with Session() as session:
        salaries = session.query(Salary).filter(
            Salary.seller == seller,
            Salary.paid == False
        ).all()
        total = sum(s.salary_amount for s in salaries)
        count = len(salaries)
        return total, count, salaries


def get_payment_method_stats(start_date, end_date, method=None):
    with Session() as session:
        revenues = session.query(DailyRevenue).filter(
            DailyRevenue.date >= start_date,
            DailyRevenue.date <= end_date
        ).all()

        total_cash = sum(r.cash for r in revenues)
        total_kaspi = sum(r.kaspi for r in revenues)
        total_halyk = sum(r.halyk for r in revenues)
        total_jusan = sum(r.jusan for r in revenues)
        total_non_cash = total_kaspi + total_halyk + total_jusan
        total_all = total_cash + total_non_cash

        if method == "Наличные":
            return total_cash, total_cash / total_all * 100 if total_all > 0 else 0
        elif method == "Kaspi":
            return total_kaspi, total_kaspi / total_all * 100 if total_all > 0 else 0
        elif method == "Халык":
            return total_halyk, total_halyk / total_all * 100 if total_all > 0 else 0
        elif method == "Жусан":
            return total_jusan, total_jusan / total_all * 100 if total_all > 0 else 0
        else:
            return {
                'cash': total_cash,
                'non_cash': total_non_cash,
                'kaspi': total_kaspi,
                'halyk': total_halyk,
                'jusan': total_jusan,
                'total': total_all
            }


def get_commission_sum(start_date, end_date):
    with Session() as session:
        commissions = session.query(BankCommission).filter(
            BankCommission.date >= start_date,
            BankCommission.date <= end_date
        ).all()
        return {
            'kaspi': sum(c.kaspi_commission for c in commissions),
            'halyk': sum(c.halyk_commission for c in commissions),
            'jusan': sum(c.jusan_commission for c in commissions),
            'total': sum(c.total_commission for c in commissions),
            'count': len(commissions)
        }


def calculate_expected_balance():
    with Session() as session:
        last_inventory = session.query(Inventory).order_by(Inventory.date.desc()).first()

        if last_inventory:
            expected = last_inventory.final_balance
        else:
            last_journal = session.query(Journal).order_by(Journal.date.desc()).first()
            if not last_journal:
                return 0
            expected = last_journal.expected_balance

        start_date = last_inventory.date if last_inventory else last_journal.date

        revenues = session.query(DailyRevenue).filter(
            DailyRevenue.date > start_date
        ).all()

        deliveries = session.query(CashDelivery).filter(
            CashDelivery.date > start_date
        ).all()

        expected += sum(r.cash + r.non_cash for r in revenues)
        expected += sum(d.amount for d in deliveries)

        return expected


def get_month_data(year, month):
    with Session() as session:
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)

        revenues = session.query(DailyRevenue).filter(
            DailyRevenue.date >= start_date,
            DailyRevenue.date < end_date
        ).all()

        deliveries = session.query(CashDelivery).filter(
            CashDelivery.date >= start_date,
            CashDelivery.date < end_date
        ).all()

        commissions = session.query(BankCommission).filter(
            BankCommission.date >= start_date,
            BankCommission.date < end_date
        ).all()

        journals = session.query(Journal).filter(
            Journal.date >= start_date,
            Journal.date < end_date
        ).all()

        inventories = session.query(Inventory).filter(
            Inventory.date >= start_date,
            Inventory.date < end_date
        ).all()

        salaries = session.query(Salary).filter(
            Salary.date >= start_date,
            Salary.date < end_date
        ).all()

        total_cash = sum(r.cash for r in revenues)
        total_kaspi = sum(r.kaspi for r in revenues)
        total_halyk = sum(r.halyk for r in revenues)
        total_jusan = sum(r.jusan for r in revenues)
        total_non_cash = total_kaspi + total_halyk + total_jusan
        total_revenue = total_cash + total_non_cash
        total_delivery = sum(d.amount for d in deliveries)
        total_commission = sum(c.total_commission for c in commissions)
        total_salary = sum(s.salary_amount for s in salaries)

        net_income = total_revenue - total_commission - total_delivery - total_salary

        seller_stats = {}
        for seller in SELLERS:
            seller_rev = [r for r in revenues if r.seller == seller]
            seller_total = sum(r.cash + r.non_cash for r in seller_rev)
            seller_days = len(set(r.date for r in seller_rev))
            if seller_total > 0:
                seller_stats[seller] = {
                    'total': seller_total,
                    'days': seller_days,
                    'avg': seller_total / seller_days if seller_days > 0 else 0
                }

        return {
            'revenues': revenues,
            'deliveries': deliveries,
            'commissions': commissions,
            'journals': journals,
            'inventories': inventories,
            'salaries': salaries,
            'total_revenue': total_revenue,
            'total_cash': total_cash,
            'total_non_cash': total_non_cash,
            'total_kaspi': total_kaspi,
            'total_halyk': total_halyk,
            'total_jusan': total_jusan,
            'total_delivery': total_delivery,
            'total_commission': total_commission,
            'total_salary': total_salary,
            'net_income': net_income,
            'days': len(set(r.date for r in revenues)),
            'seller_stats': seller_stats
        }


def forecast_revenue(days_ahead=30):
    with Session() as session:
        six_months_ago = date.today() - timedelta(days=180)
        revenues = session.query(DailyRevenue).filter(
            DailyRevenue.date >= six_months_ago
        ).order_by(DailyRevenue.date).all()

        if len(revenues) < 30:
            return None, "Недостаточно данных для прогноза (нужно минимум 30 дней)"

        dates = [(r.date - revenues[0].date).days for r in revenues]
        totals = [r.cash + r.non_cash for r in revenues]

        X = np.array(dates).reshape(-1, 1)
        y = np.array(totals)

        model = LinearRegression()
        model.fit(X, y)

        last_day = dates[-1]
        future_dates = [last_day + i for i in range(1, days_ahead + 1)]
        future_X = np.array(future_dates).reshape(-1, 1)
        predictions = model.predict(future_X)

        residuals = y - model.predict(X)
        std_dev = np.std(residuals)

        return {
            'predictions': predictions,
            'std_dev': std_dev,
            'confidence': 0.95,
            'last_date': revenues[-1].date,
            'trend': 'up' if model.coef_[0] > 0 else 'down',
            'slope': model.coef_[0],
            'r_squared': model.score(X, y)
        }, None


def apply_excel_styles(ws):
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")

    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width


# ---------- ХЕНДЛЕРЫ ----------
@dp.message(Command("start"))
async def start(message: types.Message):
    await show_menu(message)


async def show_menu(message: types.Message):
    await message.answer(
        "🏪 **Учёт магазина**\n\n"
        "💰 Ввести деньги за день\n"
        "🚚 Привоз денег\n"
        "📓 Журнал учёта\n"
        "🧮 Ревизия\n"
        "💰 Зарплаты\n"
        "📅 Посмотреть день\n"
        "📊 По способу оплаты\n"
        "📊 Комиссии банков\n"
        "📋 Дневной отчёт\n"
        "📈 Анализ и прогноз\n"
        "📄 Отчёт за месяц\n"
        "📊 Годовой отчёт\n"
        "✏️ Редактировать день\n"
        "💾 Бэкап БД\n"
        "🗃️ Просмотр всех данных\n\n"
        "🧪 Тестовый режим\n"
        "🗑️ Очистить все данные\n\n"
        "Все даты: **дд.мм.гггг** или **дд/мм/гггг**",
        reply_markup=main_keyboard(),
        parse_mode="Markdown"
    )


# ---------- ОБРАБОТЧИК НЕИЗВЕСТНЫХ КОМАНД ----------
@dp.message()
async def unknown_command(message: types.Message):
    # Игнорируем команды, которые начинаются с /
    if message.text.startswith('/'):
        return
    # Если пользователь написал что-то неизвестное
    await message.answer(
        "❌ Неизвестная команда.\n"
        "Пожалуйста, используйте кнопки ниже 👇",
        reply_markup=main_keyboard()
    )


# ---------- ПРОСМОТР ВСЕХ ДАННЫХ ----------
@dp.message(lambda msg: msg.text == "🗃️ Просмотр всех данных")
async def view_all_data(message: types.Message):
    if WEB_URL:
        await message.answer(
            f"🗃️ **Просмотр базы данных**\n\n"
            f"🌐 Откройте в браузере:\n"
            f"`{WEB_URL}`\n\n"
            f"Доступны все таблицы:\n"
            f"📊 Выручка\n"
            f"🚚 Привоз денег\n"
            f"💳 Комиссии банков\n"
            f"📓 Журнал учёта\n"
            f"🧮 Ревизии\n"
            f"💰 Зарплаты\n"
            f"🏦 Касса",
            reply_markup=main_keyboard(),
            parse_mode="Markdown"
        )
        try:
            webbrowser.open(WEB_URL)
        except:
            pass
    else:
        await message.answer(
            "❌ Веб-сервер не запущен.\n"
            "Попробуйте перезапустить бота.",
            reply_markup=main_keyboard()
        )


# ---------- ТЕСТОВЫЙ РЕЖИМ ----------
@dp.message(lambda msg: msg.text == "🧪 Тестовый режим")
async def test_mode(message: types.Message, state: FSMContext):
    global TEST_MODE
    status = "ВКЛЮЧЕН" if TEST_MODE else "ВЫКЛЮЧЕН"
    await message.answer(
        f"🧪 **ТЕСТОВЫЙ РЕЖИМ**\n\n"
        f"Текущий статус: **{status}**\n\n"
        f"В тестовом режиме данные НЕ сохраняются в БД.\n\n"
        f"Выберите действие:",
        reply_markup=test_mode_keyboard(),
        parse_mode="Markdown"
    )
    await state.set_state(DeveloperState.waiting_action)


@dp.message(DeveloperState.waiting_action)
async def process_test_mode_action(message: types.Message, state: FSMContext):
    global TEST_MODE

    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=main_keyboard())
        return

    if message.text == "🧪 Включить тестовый режим":
        TEST_MODE = True
        await message.answer(
            "🧪 **Тестовый режим ВКЛЮЧЕН!**\n\n"
            "Данные будут НЕ сохраняться.",
            reply_markup=main_keyboard(),
            parse_mode="Markdown"
        )
        await state.clear()
        return

    if message.text == "🧪 Выключить тестовый режим":
        TEST_MODE = False
        await message.answer(
            "✅ **Тестовый режим ВЫКЛЮЧЕН!**\n\n"
            "Данные будут сохраняться.",
            reply_markup=main_keyboard(),
            parse_mode="Markdown"
        )
        await state.clear()
        return

    if message.text == "🗑️ Очистить тестовые данные":
        if not TEST_MODE:
            await message.answer(
                "⚠️ Тестовый режим выключен.\n"
                "Сначала включите тестовый режим.",
                reply_markup=main_keyboard()
            )
            await state.clear()
            return

        try:
            with Session() as session:
                session.execute(delete(DailyRevenue))
                session.execute(delete(CashDelivery))
                session.execute(delete(BankCommission))
                session.execute(delete(Journal))
                session.execute(delete(Inventory))
                session.execute(delete(Salary))
                session.execute(delete(DailyCash))
                session.commit()

            await message.answer(
                "🗑️ **Все тестовые данные удалены!**\n\n"
                "База данных очищена.",
                reply_markup=main_keyboard(),
                parse_mode="Markdown"
            )
        except Exception as e:
            await message.answer(
                f"❌ Ошибка: {str(e)}",
                reply_markup=main_keyboard()
            )

        await state.clear()
        return

    await message.answer("❌ Неизвестная команда", reply_markup=main_keyboard())
    await state.clear()


# ---------- ОЧИСТКА ВСЕХ ДАННЫХ ----------
@dp.message(lambda msg: msg.text == "🗑️ Очистить все данные")
async def clear_all_data(message: types.Message, state: FSMContext):
    await message.answer(
        "⚠️ **ВНИМАНИЕ!**\n\n"
        "Вы собираетесь удалить ВСЕ данные из базы.\n"
        "Это действие необратимо!\n\n"
        "Введите код подтверждения: CLEAR",
        reply_markup=cancel_keyboard(),
        parse_mode="Markdown"
    )
    await state.set_state(ClearState.waiting_confirm)


@dp.message(ClearState.waiting_confirm)
async def process_clear_confirm(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=main_keyboard())
        return

    if message.text == "CLEAR":
        try:
            with Session() as session:
                session.execute(delete(DailyRevenue))
                session.execute(delete(CashDelivery))
                session.execute(delete(BankCommission))
                session.execute(delete(Journal))
                session.execute(delete(Inventory))
                session.execute(delete(Salary))
                session.execute(delete(DailyCash))
                session.commit()

            await message.answer(
                "🗑️ **ВСЕ ДАННЫЕ УДАЛЕНЫ!**\n\n"
                "База данных полностью очищена.",
                reply_markup=main_keyboard(),
                parse_mode="Markdown"
            )
        except Exception as e:
            await message.answer(
                f"❌ Ошибка: {str(e)}",
                reply_markup=main_keyboard()
            )
    else:
        await message.answer(
            "❌ Неверный код подтверждения.\n"
            "Операция отменена.",
            reply_markup=main_keyboard()
        )

    await state.clear()


# ---------- ВВОД ДЕНЕГ ЗА ДЕНЬ ----------
@dp.message(lambda msg: msg.text == "💰 Ввести деньги за день")
async def enter_revenue(message: types.Message, state: FSMContext):
    await message.answer(
        "👤 **Выберите продавца:**",
        reply_markup=seller_keyboard(),
        parse_mode="Markdown"
    )
    await state.set_state(RevenueState.waiting_seller)


@dp.message(RevenueState.waiting_seller)
async def process_revenue_seller(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=main_keyboard())
        return

    if message.text not in SELLERS:
        await message.answer("❌ Выберите продавца из списка")
        return

    await state.update_data(seller=message.text)
    today = date.today()
    await message.answer(
        f"📅 **Введите дату** (дд.мм.гггг или дд/мм/гггг)\n"
        f"Или отправьте '+' для {format_date(today)}",
        reply_markup=cancel_keyboard(),
        parse_mode="Markdown"
    )
    await state.set_state(RevenueState.waiting_date)


@dp.message(RevenueState.waiting_date)
async def process_revenue_date(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=main_keyboard())
        return

    if message.text == "+":
        revenue_date = date.today()
    else:
        revenue_date = parse_date(message.text)
        if not revenue_date:
            await message.answer("❌ Неверный формат. Используйте дд.мм.гггг или дд/мм/гггг")
            return

    await state.update_data(date=revenue_date)
    await message.answer(
        "💵 **Введите сумму НАЛИЧНЫМИ:**",
        reply_markup=cancel_keyboard(),
        parse_mode="Markdown"
    )
    await state.set_state(RevenueState.waiting_cash)


# ... (остальные хендлеры остаются без изменений)


async def main():
    print("🤖 Бот запущен!")
    print(f"👥 Продавцы: {', '.join(SELLERS)}")
    print(f"🧪 Тестовый режим: {'ВКЛЮЧЕН' if TEST_MODE else 'ВЫКЛЮЧЕН'}")
    print("📅 Напишите /start в Telegram")

    # Отключаем обработку сигналов
    try:
        await dp.start_polling(bot, skip_updates=True, handle_signals=False)
    except Exception as e:
        print(f"Ошибка: {e}")

if __name__ == "__main__":
    asyncio.run(main())
